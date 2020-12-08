# -*- coding: utf-8 -*-
# @Time    : 2020/12/8 22:52
# @Author  : dashenN72
"""
简单的读写excel方法
"""

import openpyxl


class ExcelUtil:
    workBook = None
    workSheet = None

    def load_excel(self, path):
        """
        加载Excel
        :param path: 需要打开的Excel的路径
        """
        self.workBook = openpyxl.load_workbook(path)

    def get_sheet_by_name(self, name):
        """
        获取sheet对象
        :param name: sheet名
        """
        self.workSheet = self.workBook[name]

    def get_sheet_by_index(self, index=0):
        """
        获取sheet对象
        :param index: sheet的索引
        """
        # 获取workBook里所有的sheet名 -> list
        sheet_names = self.workBook.get_sheet_names()
        # 根据索引获取指定sheet
        self.workSheet = self.workBook[sheet_names[index]]

    def get_cell_value(self, col, row):
        """
        获取cell的值
        :param col: 所在列
        :param row: 所在行
        """
        try:
            return self.workSheet.cell(column=col, row=row).value
        except BaseException as e:
            return None

    def get_cell_value_by_xy(self, str):
        """
        获取cell的值
        :param str: 坐标
        """
        try:
            return self.workSheet[str].value
        except BaseException as e:
            return None

    def get_sheet_rows(self):
        """
        获取最大行数
        """
        return self.workSheet.max_row

    def get_sheet_cols(self):
        """
        获取最大列数
        """
        return self.workSheet.max_column

    def write_data(self, row, col, value, path):
        """
        写入数据
        """
        try:
            self.workSheet = self.workBook.active
            self.workSheet.cell(column=col, row=row, value=value)
            self.workBook.save(path)
            return True
        except BaseException as e:
            print(e)
            return None

    def get_excel_data(self):
        """
        获取表所有数据
        :return: list
        """
        # 方式一
        data_list = tuple(self.workSheet.values)
        # 方式二
        # data_list = []
        # for i in range(self.get_sheet_rows()):
        #     data_list.append(self.get_row_value(i + 2))
        return data_list

    def get_row_value(self, row):
        """
        获取某一行的内容
        :param row: 第几行 -> str  **从1开始**
        :return: list
        """
        # 方式一
        row_list = self.get_excel_data()[row]
        # 方式二
        # row_list = []
        # for i in self.workSheet[str(row + 1)]:
        #     row_list.append(i.value)
        return row_list

    def get_col_value(self, col='A'):
        """
        获取某一列的内容
        :param col: 第几列 -> str
        :return: list
        """
        col_list = []
        for i in self.workSheet[col]:
            col_list.append(i.value)
        return col_list

    def get_row_num(self, case_id):
        """
        获取行号
        :param case_id: 用例编号
        :return:
        """
        num = 1
        col_data = self.get_col_value()
        for data in col_data:
            if case_id == data:
                return num
            num += 1
        return 0


if __name__ == '__main__':
    excelUtil = ExcelUtil()
    from op_http import http_request
    path = './data/xxx接口测试_参数校验.xlsx'
    excelUtil.load_excel(path)
    excelUtil.get_sheet_by_name("参数校验")
    rows = excelUtil.get_sheet_rows()  # 获取行数
    for row in range(1, rows):
        data = excelUtil.get_row_value(row)  # 获取行数据
        result = http_request(data[3], data[2], data[4])
        if result == 0:
            print("不支持的接口请求方法")
        elif result == 1:
            print("接口状态码错误")
        else:
            excelUtil.write_data(row=row+1, col=7, value=result, path=path)
